#!/usr/bin/env python3
"""
merge_prod_data.py — Merge ClickHouse prod-usage CSVs into bucket CSVs.

Reads three prod-data CSVs (from ~/Downloads) and adds two columns to each
bucket CSV: `prod_used` (yes/no/unknown) and `latest_prod_timestamp`.

Also writes XLSX versions with bold headers since CSV does not support
formatting.

Usage:
  python3 scripts/merge_prod_data.py
"""

import csv
import os
import re
import sys
from openpyxl import Workbook
from openpyxl.styles import Font


def to_snake(s):
    """Convert PascalCase to snake_case (e.g. BankRedirect -> bank_redirect)."""
    if not s:
        return s
    s1 = re.sub(r'(.)([A-Z][a-z]+)', r'\1_\2', s)
    return re.sub(r'([a-z0-9])([A-Z])', r'\1_\2', s1).lower()

REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DL = os.path.expanduser("~/Downloads")

# Bucket 1 features that were NOT queried (hence prod_used = 'unknown')
B1_UNQUERIED = {
    "Billing Descriptor", "L2/L3 Data Processing", "Connector Intent Metadata",
    "Connector Testing Data", "Partner Merchant Identifier", "Installments",
    "Network Transaction ID", "Partial Authorization", "Split Payments",
}

# Bucket 2 features that were NOT queried
B2_UNQUERIED = {"Payment (Decrypt Flow)"}

# Bucket 3 features NOT queried (config-level, not event-level)
B3_UNQUERIED = {
    "3DS Decision Rule Algorithm", "3DS Routing Region UAS", "Acquirer Config Map",
    "Authentication Service Eligibility", "Card Testing Guard", "Clear PAN Retries",
    "Client Session Validation", "Conditional Routing DSL", "Surcharge DSL",
    "Connector API Version Override", "Connector Onboarding Config",
    "Credentials Identifier Mapping", "CVV Collection During Payment", "Requires CVV",
    "Default Fallback Routing", "Delayed Session Response", "Dispute Polling Interval",
    "Dynamic Fields", "Eligibility Data Storage For Auth", "Extended Card BIN",
    "Extended Card Info", "Feature Metadata", "Forex/Currency Conversion",
    "Iframe Redirection", "Implicit Customer Update", "MIT With Limited Card Data",
    "Merchant Category Code", "Merchant Country Code", "Network Tokenization Credentials",
    "Outgoing Webhook Custom Headers", "Payment Response Hash",
    "Payment Update Via Client Auth", "Payout Entity Type", "Payout Priority",
    "Payout Routing Algorithm", "Payout Tracker Mapping", "PM Collect Link",
    "PM Filters CGraph", "PM Modular Service", "Poll Config", "Process Tracker Mapping",
    "Product Type", "Raw PM Details Return", "Redirect Method", "Refund Type",
    "Routing Result Source", "Session Expiry", "Split Transactions Enabled",
    "Sub-Merchants", "Tax Connector", "Use Billing As PM Billing",
    "Vault Tokenization Disable", "Webhook Config Disabled Events", "Platform Account",
    "Order Details", "FRM Routing Algorithm", "Payment Link", "Payment Manual Update",
    "Refund Manual Update", "Dynamic Routing", "Payout Link",
}


def load_b1_prod(path):
    """Returns dict[(feature, connector)] -> last_seen"""
    lookup = {}
    with open(path) as f:
        for row in csv.DictReader(f):
            lookup[(row["feature"], row["connector"].lower())] = row["last_seen"]
    return lookup


def load_b2_prod(path):
    """Returns dict[(feature, connector, pm, pmt)] -> last_seen"""
    lookup = {}
    with open(path) as f:
        for row in csv.DictReader(f):
            lookup[(row["feature"], row["connector"].lower(), row["pm"], row["pmt"])] = row["last_seen"]
    return lookup


def load_b3_prod(path):
    """Returns dict[feature] -> last_seen"""
    lookup = {}
    with open(path) as f:
        for row in csv.DictReader(f):
            lookup[row["feature"]] = row["last_seen"]
    return lookup


def write_xlsx(path, headers, rows):
    """Write an .xlsx file with bold header row."""
    wb = Workbook()
    ws = wb.active
    ws.append(headers)
    bold = Font(bold=True)
    for cell in ws[1]:
        cell.font = bold
    for row in rows:
        ws.append(row)
    # Auto-width columns based on header length (quick approximation)
    for col_idx, h in enumerate(headers, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max(len(h) + 2, 14)
    wb.save(path)


def merge_bucket_1():
    src   = os.path.join(REPO_ROOT, "bucket_1_connector_features.csv")
    prod  = load_b1_prod(os.path.join(DL, "b1_prod_data.csv"))
    csv_out  = os.path.join(REPO_ROOT, "bucket_1_connector_features.csv")
    xlsx_out = os.path.join(REPO_ROOT, "bucket_1_connector_features.xlsx")

    with open(src) as f:
        reader = csv.DictReader(f)
        base_headers = reader.fieldnames
        rows = list(reader)

    new_headers = [h for h in base_headers if h not in ("prod_used", "latest_prod_timestamp")] + ["prod_used", "latest_prod_timestamp"]

    stats = {"yes": 0, "no": 0, "unknown": 0}
    enriched = []
    for row in rows:
        feature = row["feature"]
        connector = row["connector"].lower()
        if feature in B1_UNQUERIED:
            row["prod_used"] = "unknown"
            row["latest_prod_timestamp"] = ""
            stats["unknown"] += 1
        elif (feature, connector) in prod:
            row["prod_used"] = "yes"
            row["latest_prod_timestamp"] = prod[(feature, connector)]
            stats["yes"] += 1
        else:
            row["prod_used"] = "no"
            row["latest_prod_timestamp"] = ""
            stats["no"] += 1
        enriched.append(row)

    with open(csv_out, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=new_headers, quoting=csv.QUOTE_MINIMAL)
        writer.writeheader()
        writer.writerows(enriched)

    write_xlsx(
        xlsx_out,
        new_headers,
        [[row[h] for h in new_headers] for row in enriched],
    )

    print(f"Bucket 1: {stats['yes']} yes / {stats['no']} no / {stats['unknown']} unknown  →  {csv_out}, {xlsx_out}")


def merge_bucket_2():
    src  = os.path.join(REPO_ROOT, "bucket_2_connector_pm_features.csv")
    prod = load_b2_prod(os.path.join(DL, "b2_prod_data.csv"))
    csv_out  = os.path.join(REPO_ROOT, "bucket_2_connector_pm_features.csv")
    xlsx_out = os.path.join(REPO_ROOT, "bucket_2_connector_pm_features.xlsx")

    with open(src) as f:
        reader = csv.DictReader(f)
        base_headers = reader.fieldnames
        rows = list(reader)

    new_headers = [h for h in base_headers if h not in ("prod_used", "latest_prod_timestamp")] + ["prod_used", "latest_prod_timestamp"]

    stats = {"yes": 0, "no": 0, "unknown": 0}
    enriched = []
    for row in rows:
        feature = row["feature"]
        connector = row["connector"].lower()
        # Bucket CSV uses PascalCase (Card, BankRedirect, ApplePay) — ClickHouse
        # serializes the same enums as snake_case (card, bank_redirect, apple_pay).
        pm = to_snake(row["payment_method"])
        pmt = to_snake(row["payment_method_type"])

        # Payment (Decrypt Flow) maps back to base "Payment" key since it's the same (conn,pm,pmt) combo
        # but since we can't detect decrypt specifically, mark those as unknown
        if feature in B2_UNQUERIED:
            row["prod_used"] = "unknown"
            row["latest_prod_timestamp"] = ""
            stats["unknown"] += 1
        elif (feature, connector, pm, pmt) in prod:
            row["prod_used"] = "yes"
            row["latest_prod_timestamp"] = prod[(feature, connector, pm, pmt)]
            stats["yes"] += 1
        else:
            row["prod_used"] = "no"
            row["latest_prod_timestamp"] = ""
            stats["no"] += 1
        enriched.append(row)

    with open(csv_out, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=new_headers, quoting=csv.QUOTE_MINIMAL)
        writer.writeheader()
        writer.writerows(enriched)

    write_xlsx(
        xlsx_out,
        new_headers,
        [[row[h] for h in new_headers] for row in enriched],
    )

    print(f"Bucket 2: {stats['yes']} yes / {stats['no']} no / {stats['unknown']} unknown  →  {csv_out}, {xlsx_out}")


def merge_bucket_3():
    src  = os.path.join(REPO_ROOT, "bucket_3_core_features.csv")
    csv_out  = os.path.join(REPO_ROOT, "bucket_3_core_features.csv")
    xlsx_out = os.path.join(REPO_ROOT, "bucket_3_core_features.xlsx")

    # Merge across ClickHouse + 3 Postgres sources. "Queried" = feature appears
    # in any source (so we know it was checked). "Used" = count > 0 somewhere.
    sources = [
        (os.path.join(DL, "b3_prod_data.csv"),           "call_count",    "last_seen"),
        (os.path.join(DL, "bucket_3_events.csv"),        "call_count",    "last_seen"),
        (os.path.join(DL, "bucket_3_json_features.csv"), "call_count",    "last_seen"),
        (os.path.join(DL, "business_profile_data.csv"),  "enabled_count", "last_modified"),
        (os.path.join(DL, "merchant_data.csv"),          "enabled_count", "last_modified"),
        (os.path.join(DL, "config_data.csv"),            "enabled_count", "last_modified"),
    ]

    # Features explicitly queried by our SQL files. Even if they return 0 rows
    # (and get excluded from the exported CSV), we still know they were checked.
    KNOWN_QUERIED = {
        # bucket_3_events.sql
        "Routing Algorithm", "External 3DS Authentication", "Dispute Management",
        "FRM (Fraud Risk Management)", "Webhook Details", "External Vault",
        "Payout Type", "Payout Auto Fulfill", "Payout Recurring", "Payout Link",
        "Save Card Flow", "Off Session Payments", "Payment Link",
        "Mandate Management", "Multiple Capture", "Browser Info Collection",
        "Connector Metadata", "Customer Acceptance", "Connector Agnostic MIT",
        # bucket_3_api_events.sql
        "Health Check", "Payment Sync", "Void/Cancel Payment",
        "SDK Client Token Generation", "Auto Retries", "Manual Retry",
        "Customer Management", "Payment Method Operations", "MCA Management",
        "Merchant Account Management", "Business Profile Management",
        "Organization Management", "Card Issuer Management",
        "Subscription Management", "Blocklist", "OIDC Authentication",
        "Gateway Status Map (GSM)", "Reconciliation", "Routing Evaluate",
        "Relay Operations",
        # bucket_3_json_features.sql
        "Payment Manual Update", "Refund Manual Update", "Order Details",
        "Feature Metadata", "Payout Entity Type", "Payout Priority", "Refund Type",
    }

    queried = set(KNOWN_QUERIED)
    used = {}   # feature -> latest timestamp (may be "" if no timestamp column)
    for path, count_col, time_col in sources:
        if not os.path.exists(path):
            continue
        with open(path) as f:
            for row in csv.DictReader(f):
                # business_profile has "(config)" suffixes to disambiguate from
                # CH-side versions of the same feature — strip for matching.
                feature = row["feature"].replace(" (config)", "")
                queried.add(feature)
                try:
                    count = int(row[count_col] or 0)
                except (ValueError, KeyError):
                    count = 0
                if count > 0:
                    ts = (row.get(time_col) or "").strip()
                    if feature not in used or (ts and ts > used[feature]):
                        used[feature] = ts

    with open(src) as f:
        reader = csv.DictReader(f)
        base_headers = reader.fieldnames
        rows = list(reader)

    new_headers = [h for h in base_headers if h not in ("prod_used", "latest_prod_timestamp")] + ["prod_used", "latest_prod_timestamp"]

    stats = {"yes": 0, "no": 0, "unknown": 0}
    enriched = []
    for row in rows:
        feature = row["feature"]
        if feature in used:
            row["prod_used"] = "yes"
            row["latest_prod_timestamp"] = used[feature]
            stats["yes"] += 1
        elif feature in queried:
            row["prod_used"] = "no"
            row["latest_prod_timestamp"] = ""
            stats["no"] += 1
        else:
            row["prod_used"] = "unknown"
            row["latest_prod_timestamp"] = ""
            stats["unknown"] += 1
        enriched.append(row)

    with open(csv_out, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=new_headers, quoting=csv.QUOTE_MINIMAL)
        writer.writeheader()
        writer.writerows(enriched)

    write_xlsx(
        xlsx_out,
        new_headers,
        [[row[h] for h in new_headers] for row in enriched],
    )

    print(f"Bucket 3: {stats['yes']} yes / {stats['no']} no / {stats['unknown']} unknown  →  {csv_out}, {xlsx_out}")


def regenerate_report():
    """Rewrite feature_extraction_report.csv with prod_used columns added."""
    report_path = os.path.join(REPO_ROOT, "feature_extraction_report.csv")

    bucket_files = {
        "Bucket 1": ("bucket_1_connector_features.csv", "Connector × Feature (payload-level, PM-agnostic)"),
        "Bucket 2": ("bucket_2_connector_pm_features.csv", "Connector × PM × PMT × Feature"),
        "Bucket 3": ("bucket_3_core_features.csv", "Core features (connector-agnostic)"),
    }

    stats = {}
    for bucket, (fname, desc) in bucket_files.items():
        path = os.path.join(REPO_ROOT, fname)
        with open(path) as f:
            rows = list(csv.DictReader(f))
        total = len(rows)
        cy_c = {"covered": 0, "not_covered": 0, "no_cypress_config": 0}
        pr_c = {"yes": 0, "no": 0, "unknown": 0}
        for row in rows:
            cy_c[row.get("cypress_test_status", "")] = cy_c.get(row.get("cypress_test_status", ""), 0) + 1
            pr_c[row["prod_used"]] = pr_c.get(row["prod_used"], 0) + 1
        cy_pct = f"{round(cy_c.get('covered', 0) / total * 100)}%" if total else "0%"
        # Prod coverage % = yes / (yes + no), excluding unknown from the denominator
        denom = pr_c["yes"] + pr_c["no"]
        pr_pct = f"{round(pr_c['yes'] / denom * 100)}%" if denom else "0%"
        stats[bucket] = {
            "description": desc, "total": total,
            "cy_covered": cy_c.get("covered", 0),
            "cy_not_covered": cy_c.get("not_covered", 0),
            "cy_no_config": cy_c.get("no_cypress_config", 0),
            "cy_pct": cy_pct,
            "prod_yes": pr_c["yes"], "prod_no": pr_c["no"], "prod_unknown": pr_c["unknown"],
            "prod_pct": pr_pct,
        }

    total_row = {"description": "All buckets combined",
                 "total": sum(s["total"] for s in stats.values()),
                 "cy_covered":     sum(s["cy_covered"] for s in stats.values()),
                 "cy_not_covered": sum(s["cy_not_covered"] for s in stats.values()),
                 "cy_no_config":   sum(s["cy_no_config"] for s in stats.values()),
                 "prod_yes":       sum(s["prod_yes"] for s in stats.values()),
                 "prod_no":        sum(s["prod_no"] for s in stats.values()),
                 "prod_unknown":   sum(s["prod_unknown"] for s in stats.values())}
    total_row["cy_pct"]   = f"{round(total_row['cy_covered'] / total_row['total'] * 100)}%" if total_row["total"] else "0%"
    denom = total_row["prod_yes"] + total_row["prod_no"]
    total_row["prod_pct"] = f"{round(total_row['prod_yes'] / denom * 100)}%" if denom else "0%"

    headers = [
        "bucket", "description", "total_rows",
        "cypress_covered", "cypress_not_covered", "cypress_no_config", "cypress_coverage_pct",
        "prod_used_yes", "prod_used_no", "prod_used_unknown", "prod_usage_pct",
    ]

    with open(report_path, "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(headers)
        for bucket, s in stats.items():
            w.writerow([bucket, s["description"], s["total"],
                        s["cy_covered"], s["cy_not_covered"], s["cy_no_config"], s["cy_pct"],
                        s["prod_yes"], s["prod_no"], s["prod_unknown"], s["prod_pct"]])
        s = total_row
        w.writerow(["TOTAL", s["description"], s["total"],
                    s["cy_covered"], s["cy_not_covered"], s["cy_no_config"], s["cy_pct"],
                    s["prod_yes"], s["prod_no"], s["prod_unknown"], s["prod_pct"]])

    # XLSX version with bold header
    xlsx_path = report_path.replace(".csv", ".xlsx")
    with open(report_path) as f:
        rows = list(csv.reader(f))
    write_xlsx(xlsx_path, rows[0], rows[1:])
    print(f"Report:   {report_path}, {xlsx_path}")


def main():
    for f in ["b1_prod_data.csv", "b2_prod_data.csv", "b3_prod_data.csv"]:
        p = os.path.join(DL, f)
        if not os.path.exists(p):
            print(f"ERROR: missing {p}", file=sys.stderr)
            sys.exit(1)

    merge_bucket_1()
    merge_bucket_2()
    merge_bucket_3()
    regenerate_report()


if __name__ == "__main__":
    main()
